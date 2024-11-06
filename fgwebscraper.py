# webscraping and download data
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys

current_date = datetime.now().strftime('%m%d')
##### LOAD WEBDRIVER TO BEGIN BUILDING WEBSCRAPER FOR FANGRAPHS #####
chrome_options = Options()
download_dir = r"C:\path\to\directory"

prefs = {"download.default_directory": download_dir,
         "download.prompt_for_download": False,
         "download.directory_upgrade": True,
         "safebrowsing.enabled": True}

chrome_options.add_experimental_option("prefs", prefs)

# setup chrome webdriver
driver_path = r"C:\path\to\directory\with\driver"
driver = webdriver.Chrome(executable_path=driver_path, options = chrome_options)

# open fangraphs sign in page
driver.get('https://blogs.fangraphs.com/wp-login.php?redirect_to=https://www.fangraphs.com/')

# locate username and password field
username_field = driver.find_element('id', 'user_login')
password_field = driver.find_element('id', 'user_pass')

username_field.send_keys('username')
password_field.send_keys('password')

password_field.send_keys(Keys.RETURN)
# wait for login to complete
time.sleep(5)

# List URL's to navigate to
urls = {'Batters - Full Season': 'https://www.fangraphs.com/leaders/major-league?pos=all&stats=bat&lg=all&season=2024&season1=2024&ind=0&type=c%2C2%2C3%2C4%2C5%2C6%2C7%2C8%2C9%2C10%2C11%2C12%2C13%2C14%2C16%2C21%2C22%2C23%2C34%2C35%2C37%2C38%2C39%2C40%2C43%2C44%2C45%2C47%2C211%2C308%2C311&month=33&qual=50&v_cr=202301',
        'Batters - Last 30': 'https://www.fangraphs.com/leaders/major-league?pos=all&stats=bat&lg=all&season=2024&season1=2024&ind=0&v_cr=202301&type=c%2C2%2C3%2C4%2C5%2C6%2C7%2C8%2C9%2C10%2C11%2C12%2C13%2C14%2C16%2C21%2C22%2C23%2C34%2C35%2C37%2C38%2C39%2C40%2C43%2C44%2C45%2C47%2C211%2C308%2C311&month=3&qual=30',
        'Batters - Last 14': 'https://www.fangraphs.com/leaders/major-league?pos=all&stats=bat&lg=all&season=2024&season1=2024&ind=0&v_cr=202301&type=c%2C2%2C3%2C4%2C5%2C6%2C7%2C8%2C9%2C10%2C11%2C12%2C13%2C14%2C16%2C21%2C22%2C23%2C34%2C35%2C37%2C38%2C39%2C40%2C43%2C44%2C45%2C47%2C211%2C308%2C311&month=2&qual=20',
        'Batters - Last 7': 'https://www.fangraphs.com/leaders/major-league?pos=all&stats=bat&lg=all&season=2024&season1=2024&ind=0&v_cr=202301&type=c%2C2%2C3%2C4%2C5%2C6%2C7%2C8%2C9%2C10%2C11%2C12%2C13%2C14%2C16%2C21%2C22%2C23%2C34%2C35%2C37%2C38%2C39%2C40%2C43%2C44%2C45%2C47%2C211%2C308%2C311&month=1&qual=10',
        'Pitchers - Full Season': 'https://www.fangraphs.com/leaders/major-league?pos=all&lg=all&season=2024&season1=2024&ind=0&sortcol=14&sortdir=desc&stats=sta&v_cr=202301&type=c%2C4%2C5%2C7%2C8%2C13%2C6%2C45%2C15%2C18%2C47%2C48%2C49%2C51%2C120%2C121%2C329%2C324%2C325%2C327%2C328&month=33&qual=30',
        'Pitchers - Last 30': 'https://www.fangraphs.com/leaders/major-league?pos=all&lg=all&season=2024&season1=2024&ind=0&sortcol=14&sortdir=desc&stats=sta&v_cr=202301&type=c%2C4%2C5%2C7%2C8%2C13%2C6%2C45%2C15%2C18%2C47%2C48%2C49%2C51%2C120%2C121%2C329%2C324%2C325%2C327%2C328&month=3&qual=20',
        'Pitchers - Last 14': 'https://www.fangraphs.com/leaders/major-league?pos=all&lg=all&season=2024&season1=2024&ind=0&sortcol=14&sortdir=desc&stats=sta&v_cr=202301&type=c%2C4%2C5%2C7%2C8%2C13%2C6%2C45%2C15%2C18%2C47%2C48%2C49%2C51%2C120%2C121%2C329%2C324%2C325%2C327%2C328&month=2&qual=5',
        'Pitchers - Last 7': 'https://www.fangraphs.com/leaders/major-league?pos=all&lg=all&season=2024&season1=2024&ind=0&sortcol=14&sortdir=desc&stats=sta&v_cr=202301&type=c%2C4%2C5%2C7%2C8%2C13%2C6%2C45%2C15%2C18%2C47%2C48%2C49%2C51%2C120%2C121%2C329%2C324%2C325%2C327%2C328&month=1&qual=1',
        'Probable Starters': 'https://www.fangraphs.com/leaders/major-league?pos=all&stats=pit&lg=all&qual=0&type=8&season=2024&month=0&season1=2024&ind=0&team=0&rost=0&age=0&filter=&players=p2024-09-18'
        }



##### CREATE EXCEL WRITER OBJECT TO POPULATE WORKBOOK WITH DATA #####
# create excel writer object
excel_path = os.path.join(download_dir, f'Fangraphs_Stats{current_date}.xlsx')
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    
    # Iterate over each key-value pair in the dictionary
    for sheet_name, url in urls.items():
        # navigate to the url
        driver.get(url)
        time.sleep(3)
        
        if sheet_name == 'Probable Starters':
            
            scores_tab = driver.find_element(By.XPATH, "//div[@class='menu-item-label' and text()='Scores']")
            
            actions = ActionChains(driver)
            actions.move_to_element(scores_tab).perform()
            time.sleep(3)
            
            probable_pitchers_link = WebDriverWait(driver, 10).until(
                EC.visibility_of_elemeent_located((By.XPATH, "//a[contains(text(), 'Probable Pitchers')]"))
            )
            probable_pitchers_link.click()
            
        else:
            # Click the 'Export Data' button to download
            export = driver.find_element(By.XPATH, "//a[contains(text(), 'Export Data')]")
            export.click()
            time.sleep(3)
        
        # Find the latest file in the download directory
        list_of_files = os.listdir(download_dir)
        latest_file = max([os.path.join(download_dir, f) for f in list_of_files], key=os.path.getctime)
        
        # load the downloaded csv file into a dataframe
        df = pd.read_csv(latest_file)
        
        # Save the DataFrame to the excel file u nder the current sheet name
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        os.remove(latest_file)
        


driver.quit()

print(f"All datasets have been saved to {excel_path}")


# Create function that drops the last 3 columns from each sheet, converts specified columns to percentages, and changed number of places after decimal for all values
def format_workbook(workbook, formatting_rules, decimal_places=3):
    """
    Drops the last three columns, converts specified columns to percentages,
    and adjusts the number of decimal places for all values in each sheet of the workbook.

    Parameters:
    workbook (openpyxl.workbook.Workbook): The workbook to process.
    formatting_rules (dict): Dictionary where keys are sheet names and values are 
                             dictionaries with 'percentage_columns' and 'decimal_columns' lists.
    decimal_places (int): Number of decimal places to apply to columns in decimal_columns.
    """
    for sheet_name, rules in formatting_rules.items():
        ws = workbook[sheet_name]

        # Drop the last three columns
        for _ in range(3):
            total_columns = ws.max_column
            ws.delete_cols(total_columns)

        # Format specified columns as percentages
        for col_letter in rules.get('percentage_columns', []):
            for cell in ws[col_letter]:
                cell.number_format = '0.00%'

        # Adjust the number of decimal places for specified columns
        decimal_format = f'0.{"0"*decimal_places}'
        for col_letter in rules.get('decimal_columns', []):
            for cell in ws[col_letter]:
                cell.number_format = decimal_format                    
                    
# Create variables to store arguments for function
# Define columns for formatting as percentages and with specified decimal places
formatting_rules = {
    'Batters - Full Season': {'percentage_columns': ['T', 'U', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF'], 'decimal_columns': ['S', 'V', 'W', 'X', 'Y']},
    'Batters - Last 30': {'percentage_columns': ['T', 'U', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF'], 'decimal_columns': ['S', 'V', 'W', 'X', 'Y']},
    'Batters - Last 14': {'percentage_columns': ['T', 'U', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF'], 'decimal_columns': ['S', 'V', 'W', 'X', 'Y']},
    'Batters - Last 7': {'percentage_columns': ['T', 'U', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF'], 'decimal_columns': ['S', 'V', 'W', 'X', 'Y']},
    'Pitchers - Full Season': {'percentage_columns': ['L', 'M', 'N', 'O', 'P', 'Q', 'T', 'V'], 'decimal_columns': ['H', 'I']},
    'Pitchers - Last 30': {'percentage_columns': ['L', 'M', 'N', 'O', 'P', 'Q', 'T', 'V'], 'decimal_columns': ['H', 'I']},
    'Pitchers - Last 14': {'percentage_columns': ['L', 'M', 'N', 'O', 'P', 'Q', 'T', 'V'], 'decimal_columns': ['H', 'I']},
    'Pitchers - Last 7': {'percentage_columns': ['L', 'M', 'N', 'O', 'P', 'Q', 'T', 'V'], 'decimal_columns': ['H', 'I']}
    }
      

def convert_to_table_and_autofit(workbook):
    """
    Converts the data in each sheet of the workbook into a table and auto-fits the columns.

    Parameters:
    workbook (openpyxl.workbook.Workbook): The workbook to process.
    """
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        # Define the table range
        table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        
        # Create a table object
        table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=table_range)
        
        # Apply a table style
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        
        # Add the table to the worksheet
        ws.add_table(table)

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            
            for cell in col:
                try:
                    # Calculate the length of the cell's content
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set the column width to auto-fit (similar to "Alt + H + O + I")
            adjusted_width = max_length + 8 if max_length > 0 else 8  # 10 as a fallback width
            ws.column_dimensions[column].width = adjusted_width            

# Load the workbook
workbook = load_workbook(excel_path)

# Apply the formatting functions
format_workbook(workbook, formatting_rules, decimal_places=3)
convert_to_table_and_autofit(workbook)

# Save the changes to the workbook
workbook.save(excel_path)
